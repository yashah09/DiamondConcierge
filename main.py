from flask import Flask, request, jsonify
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
import os
import io
import uuid
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# CONFIGURATION
FOLDER_ID = '1diAVIuJdsOQhLEQuFzie6QACakeOie25'
INVENTORY_FILE_ID = '1McHVVICDeeMRiA1fRU7inHmbSUCzeOD2'
MAKE_WEBHOOK_URL = 'https://hook.eu2.make.com/h6jsruunr7u01wobm995dj8wtcmafph8'
SCOPES = ['https://www.googleapis.com/auth/drive']

creds = service_account.Credentials.from_service_account_file(
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"], scopes=SCOPES
)
drive_service = build('drive', 'v3', credentials=creds)

app = Flask(__name__)

def get_latest_inventory_from_drive():
    try:
        request = drive_service.files().get_media(fileId=INVENTORY_FILE_ID)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return pd.read_excel(fh, engine="openpyxl")
    except Exception as e:
        print("Error reading inventory:", e)
        return None

def filter_inventory(df, filters):
    df = df.copy()

    shape_map = {
        "CU": "Cushion", "CB": "Cushion", "AS": "Asscher", "SQEM": "Asscher",
        "RD": "Round", "PS": "Pear", "OV": "Oval", "EM": "Emerald", "PR": "Princess"
    }
    if "shape" in filters:
        target = shape_map.get(filters["shape"].upper(), filters["shape"])
        df = df[df["Shape"].fillna('').str.upper() == target.upper()]

    if "certified" in filters:
        df = df[df["Lab Name"].notna()] if filters["certified"] else df[df["Lab Name"].isna()]

    if "size_min" in filters and "size_max" in filters:
        df = df[(df["Cts"] >= filters["size_min"]) & (df["Cts"] <= filters["size_max"])]

    if "color_min" in filters and "color_max" in filters:
        color_order = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
        cmin = color_order.index(filters["color_min"].upper())
        cmax = color_order.index(filters["color_max"].upper())
        df = df[df["Color"].fillna('').str.upper().isin(color_order[cmin:cmax+1])]

    if "clarity_min" in filters and "clarity_max" in filters:
        clarity_order = ["IF", "VVS1", "VVS2", "VS1", "VS2", "SI1", "SI2", "SI3", "I1", "I2"]
        clmin = clarity_order.index(filters["clarity_min"].upper())
        clmax = clarity_order.index(filters["clarity_max"].upper())
        df = df[df["Clarity"].fillna('').str.upper().isin(clarity_order[clmin:clmax+1])]

    for key, col in [("cut", "Cut"), ("polish", "Pol"), ("symmetry", "Sym")]:
        if key in filters:
            df = df[df[col].fillna('').str.upper().isin([v.upper() for v in filters[key]])]

    if "fluorescence" in filters:
        flour_map = {
            "NON": ["NONE", "NON"],
            "FNT": ["FAINT", "FNT"],
            "MED": ["MEDIUM", "MED"],
            "STR": ["STRONG", "STR", "STG"],
            "VST": ["VERY STRONG", "VST"]
        }
        valid_vals = []
        for code in filters["fluorescence"]:
            for key, aliases in flour_map.items():
                if code.upper() in aliases or code.upper() == key:
                    valid_vals.extend(aliases)
        df = df[df["Fluo."].fillna('').str.upper().isin([v.upper() for v in valid_vals])]

    if "discount_max" in filters:
        df = df[df["Discount"] <= filters["discount_max"]]
    if "discount_min" in filters:
        df = df[df["Discount"] >= filters["discount_min"]]

    if "ppc_max" in filters:
        df = df[df["PPC"] <= filters["ppc_max"]]
    if "ppc_min" in filters:
        df = df[df["PPC"] >= filters["ppc_min"]]

    if "td_min" in filters and "td_max" in filters:
        df = df[(df["Total Depth"] >= filters["td_min"]) & (df["Total Depth"] <= filters["td_max"])]
    if "table_min" in filters and "table_max" in filters:
        df = df[(df["Table Size"] >= filters["table_min"]) & (df["Table Size"] <= filters["table_max"])]

    return df.reset_index(drop=True)

def summarize(df):
    return {
        "stones": int(df.shape[0]),
        "carats": round(df["Cts"].sum(), 2),
        "ppc_avg": round(df["PPC"].mean(), 2),
        "disc_avg": round(df["Discount"].mean(), 2),
        "rap_avg": round(df["RAP Price"].mean(), 2),
        "total_value": round(df["Total Value"].sum(), 2)
    }

def beautify_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    ws.insert_rows(1, amount=2)

    red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    pink_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True, name="Aptos Narrow", size=11)
    black_font = Font(color="000000", name="Aptos Narrow", size=11)

    headers = ["", "Stones", "Carats", "", "Rap Average", "PPC Average", "Avg Disc", "", "", "", "Total Value"]
    formulas = [
        "Selection", "=SUBTOTAL(3,B4:B3153)", "=SUBTOTAL(9,E4:E3153)", "",
        "=SUBTOTAL(9,M4:M3153)/H2", "=SUBTOTAL(9,P4:P3153)/H2", "=((K2/J2)-1)*100",
        "", "", "", "=IF(G2<200,SUBTOTAL(9,P4:P3153),0)"
    ]

    for i in range(11):
        col = 6 + i
        ws.cell(row=1, column=col, value=headers[i])
        ws.cell(row=1, column=col).fill = red_fill
        ws.cell(row=1, column=col).font = white_bold

        cell = ws.cell(row=2, column=col)
        cell.fill = pink_fill
        if formulas[i]:
            cell.value = formulas[i]
            if get_column_letter(col) == 'H':
                cell.number_format = '0.00'
            elif get_column_letter(col) in ['J', 'K']:
                cell.number_format = '0'
        cell.font = Font(bold=(col == 6), name="Aptos Narrow", size=11, color="000000")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = red_fill
        cell.font = Font(color="FFFFFF", name="Aptos Narrow", size=11, bold=False)

    # Add borders to F1:P2 and A3:AF4000
from openpyxl.styles.borders import Border, Side
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# F1:P2 borders
for row in range(1, 3):
    for col in range(6, 17):
        ws.cell(row=row, column=col).border = border_style

# A3:AF4000 borders
for row in range(3, 4001):
    for col in range(1, 33):
        ws.cell(row=row, column=col).border = border_style

# Add borders to F1:P2 and A3:AF4000
from openpyxl.styles.borders import Border, Side
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# F1:P2 borders
for row in range(1, 3):
    for col in range(6, 17):
        ws.cell(row=row, column=col).border = border_style

# A3:AF4000 borders
for row in range(3, 4001):
    for col in range(1, 33):
        ws.cell(row=row, column=col).border = border_style

# Number formats for F2:P2
ws['H2'].number_format = '0.00'
ws['J2'].number_format = '0'
ws['K2'].number_format = '0'
ws['L2'].number_format = '0.00'
ws['P2'].number_format = '0'

# Data columns (E4:E4000 = carats), N4:N4000 = Discount %, O4:O4000 = PPC, P4:P4000 = Total Value
for row in range(4, 4001):
    ws[f'E{row}'].number_format = '0.00'
    ws[f'N{row}'].number_format = '0.00'
    ws[f'O{row}'].number_format = '0'
    ws[f'P{row}'].number_format = '0'

wb.save(path)

@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    email = data.get("email")
    filters = data.get("filters", {})

    if not email:
        return jsonify({"error": "Missing email"}), 400

    df = get_latest_inventory_from_drive()
    if df is None:
        return jsonify({"error": "Could not load inventory"}), 500

    filtered_df = filter_inventory(df, filters)
    if filtered_df.empty:
        return jsonify({"error": "No matching stones"}), 404

    summary = summarize(filtered_df)
    filename = f"filtered_{uuid.uuid4().hex[:6]}.xlsx"
    filtered_df.to_excel(filename, index=False)
    beautify_excel(filename)

    metadata = {"name": filename, "parents": [FOLDER_ID]}
    media = MediaFileUpload(filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file = drive_service.files().create(body=metadata, media_body=media, fields='id').execute()
    file_id = file.get("id")
    drive_service.permissions().create(fileId=file_id, body={"role": "reader", "type": "anyone"}).execute()

    link = f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"
    os.remove(filename)

    try:
        requests.post(MAKE_WEBHOOK_URL, json={"email": email, "file_link": link, "file_name": filename})
    except:
        pass

    return jsonify({"message": "File filtered", "summary": summary, "link": link})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 10000)))
